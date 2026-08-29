from __future__ import annotations

from collections.abc import Mapping, Sequence
from pathlib import Path
import math
import re

import numpy as np
from openpyxl import load_workbook

from gf.dl.adapters.base import AdapterError
from gf.dl.contracts import UnifiedSample


DATASET_ID = "xylene_e_nose"
STANDARD_SENSOR_IDS = ("HKUST1", "CuBDC", "CuBPDC", "UiO66", "UiO67", "UiO68-NH2")
CHANNEL_ALIASES = {
    "HKUST1": "HKUST1",
    "HKUST1a": "HKUST1",
    "CuBDC": "CuBDC",
    "CuBDCa": "CuBDC",
    "CuBPDCa": "CuBPDC",
    "UiO66": "UiO66",
    "UiO66a": "UiO66",
    "UiO67a": "UiO67",
    "UiO-68NH2a": "UiO68-NH2",
    "UiO68-NH2": "UiO68-NH2",
}
LABEL_PATTERN = re.compile(r"^(?P<total>\d+(?:\.\d+)?) ppm (?P<body>.+)$")
PURE_PATTERN = re.compile(r"^(?P<component>[mop])-xylene$")
MIXTURE_PATTERN = re.compile(r"^(?P<components>[mop](?:-[mop])+?)=(?P<ratios>\d+(?:-\d+)*)$")


class XyleneENoseAdapter:
    dataset_id = DATASET_ID

    def __init__(self, *, dataset_root: Path, workbooks: Sequence[Mapping[str, object]], window_rows: int) -> None:
        self._dataset_root = dataset_root.resolve()
        if not self._dataset_root.is_dir():
            raise AdapterError(f"xylene dataset root does not exist: {self._dataset_root}")
        if not workbooks:
            raise AdapterError("xylene smoke config must contain at least one workbook")
        if window_rows <= 0:
            raise AdapterError("window_rows must be positive")
        self._workbooks = tuple(dict(workbook) for workbook in workbooks)
        self._window_rows = window_rows

    @classmethod
    def from_config(cls, config: Mapping[str, object], *, project_root: Path) -> "XyleneENoseAdapter":
        if config.get("dataset_id") != cls.dataset_id:
            raise AdapterError(f"expected dataset_id {cls.dataset_id!r}, got {config.get('dataset_id')!r}")
        dataset_root = config.get("dataset_root")
        workbooks = config.get("workbooks")
        if not isinstance(dataset_root, str) or not dataset_root:
            raise AdapterError("dataset_root must be a non-empty string")
        if not isinstance(workbooks, list):
            raise AdapterError("workbooks must be a list")
        try:
            window_rows = int(config["window_rows"])
        except (KeyError, TypeError, ValueError) as exc:
            raise AdapterError("window_rows must be a valid integer") from exc
        return cls(
            dataset_root=(project_root / dataset_root),
            workbooks=workbooks,
            window_rows=window_rows,
        )

    def load_samples(self) -> list[UnifiedSample]:
        samples: list[UnifiedSample] = []
        seen_names: set[str] = set()
        for workbook_spec in self._workbooks:
            filename = workbook_spec.get("file")
            split = workbook_spec.get("split")
            if not isinstance(filename, str) or not filename:
                raise AdapterError(f"workbook file must be a non-empty string, got {filename!r}")
            if split not in {"train", "val", "test"}:
                raise AdapterError(f"invalid split for workbook {filename!r}: {split!r}")
            if filename in seen_names:
                raise AdapterError(f"duplicate workbook {filename!r}")
            seen_names.add(filename)
            samples.append(self._load_workbook(filename, str(split)))
        return samples

    def _load_workbook(self, filename: str, split: str) -> UnifiedSample:
        path = (self._dataset_root / filename).resolve()
        if not path.is_relative_to(self._dataset_root):
            raise AdapterError(f"workbook path escapes dataset root: {filename!r}")
        if not path.is_file():
            raise AdapterError(f"workbook does not exist: {path}")
        target, nominal_total, family = parse_xylene_label(path.name)

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            header_rows = list(worksheet.iter_rows(min_row=1, max_row=2, values_only=True))
            if len(header_rows) != 2 or header_rows[1][0] != "Time":
                raise AdapterError(f"unexpected two-row header in {path.name!r}")
            column_by_sensor = _resolve_sensor_columns(header_rows[1], path.name)

            times_min: list[float] = []
            values_by_sensor: dict[str, list[float]] = {sensor_id: [] for sensor_id in STANDARD_SENSOR_IDS}
            masks_by_sensor: dict[str, list[bool]] = {sensor_id: [] for sensor_id in STANDARD_SENSOR_IDS}
            for row in worksheet.iter_rows(min_row=3, values_only=True):
                if len(times_min) >= self._window_rows:
                    break
                time_value = _finite_float(row[0], field="Time", filename=path.name)
                times_min.append(time_value)
                for sensor_id in STANDARD_SENSOR_IDS:
                    column_index = column_by_sensor[sensor_id]
                    raw_value = row[column_index]
                    if raw_value is None:
                        values_by_sensor[sensor_id].append(0.0)
                        masks_by_sensor[sensor_id].append(False)
                        continue
                    value = _finite_float(raw_value, field=sensor_id, filename=path.name)
                    values_by_sensor[sensor_id].append(value)
                    masks_by_sensor[sensor_id].append(True)
            if len(times_min) != self._window_rows:
                raise AdapterError(
                    f"workbook {path.name!r} yielded {len(times_min)} rows, expected {self._window_rows}"
                )
        finally:
            workbook.close()

        time_s = np.asarray(times_min, dtype=np.float64) * 60.0
        if time_s.size > 1 and np.any(np.diff(time_s) <= 0.0):
            raise AdapterError(f"Time must be strictly increasing in selected window of {path.name!r}")

        signals: list[np.ndarray] = []
        valid_masks: list[np.ndarray] = []
        qualities: list[np.ndarray] = []
        times: list[np.ndarray] = []
        for sensor_id in STANDARD_SENSOR_IDS:
            signal = np.asarray(values_by_sensor[sensor_id], dtype=np.float32).reshape(-1, 1)
            mask = np.asarray(masks_by_sensor[sensor_id], dtype=np.bool_).reshape(-1, 1)
            signals.append(signal)
            valid_masks.append(mask)
            qualities.append(mask[:, 0].astype(np.float32))
            times.append(time_s)

        return UnifiedSample(
            signals=tuple(signals),
            sensor_id=STANDARD_SENSOR_IDS,
            sensor_type=("qcm_frequency_shift",) * len(STANDARD_SENSOR_IDS),
            valid_mask=tuple(valid_masks),
            quality=tuple(qualities),
            time=tuple(times),
            target=target,
            target_mask=np.ones(3, dtype=np.bool_),
            group_id=path.name,
            dataset_id=self.dataset_id,
            metadata={
                "workbook_name": path.name,
                "split": split,
                "nominal_total_ppm": nominal_total,
                "composition_family": family,
            },
        )


def parse_xylene_label(filename: str) -> tuple[np.ndarray, float, str]:
    stem = Path(filename).stem
    if stem.startswith("max concentration"):
        raise AdapterError(f"near-saturation workbook has no quantitative ppm label: {filename!r}")
    match = LABEL_PATTERN.fullmatch(stem)
    if match is None:
        raise AdapterError(f"cannot parse xylene label from filename {filename!r}")
    total = float(match.group("total"))
    body = match.group("body")
    target_by_component = {"m": 0.0, "o": 0.0, "p": 0.0}

    pure_match = PURE_PATTERN.fullmatch(body)
    if pure_match is not None:
        target_by_component[pure_match.group("component")] = total
        family = "pure"
    else:
        mixture_match = MIXTURE_PATTERN.fullmatch(body)
        if mixture_match is None:
            raise AdapterError(f"cannot parse xylene mixture ratios from filename {filename!r}")
        components = mixture_match.group("components").split("-")
        ratios = [int(value) for value in mixture_match.group("ratios").split("-")]
        if len(components) != len(ratios) or len(set(components)) != len(components):
            raise AdapterError(f"component and ratio structure is invalid in filename {filename!r}")
        if any(ratio <= 0 for ratio in ratios):
            raise AdapterError(f"mixture ratios must be positive in filename {filename!r}")
        ratio_sum = sum(ratios)
        for component, ratio in zip(components, ratios, strict=True):
            target_by_component[component] = total * ratio / ratio_sum
        family = "binary" if len(components) == 2 else "ternary"

    target = np.array(
        [target_by_component["m"], target_by_component["o"], target_by_component["p"]],
        dtype=np.float32,
    )
    if not math.isclose(float(target.sum()), total, rel_tol=0.0, abs_tol=1e-5):
        raise AdapterError(f"parsed component labels do not sum to total ppm in {filename!r}")
    return target, total, family


def _resolve_sensor_columns(header: tuple[object, ...], filename: str) -> dict[str, int]:
    columns: dict[str, int] = {}
    for column_index, raw_name in enumerate(header[1:], start=1):
        if raw_name is None:
            continue
        normalized = CHANNEL_ALIASES.get(str(raw_name))
        if normalized is None:
            raise AdapterError(f"unknown QCM channel {raw_name!r} in {filename!r}")
        if normalized in columns:
            raise AdapterError(f"duplicate normalized QCM channel {normalized!r} in {filename!r}")
        columns[normalized] = column_index
    missing = set(STANDARD_SENSOR_IDS) - set(columns)
    extra = set(columns) - set(STANDARD_SENSOR_IDS)
    if missing or extra:
        raise AdapterError(f"QCM channel contract mismatch in {filename!r}: missing={sorted(missing)}, extra={sorted(extra)}")
    return columns


def _finite_float(value: object, *, field: str, filename: str) -> float:
    try:
        numeric = float(value)
    except (TypeError, ValueError) as exc:
        raise AdapterError(f"non-numeric {field} value {value!r} in {filename!r}") from exc
    if not math.isfinite(numeric):
        raise AdapterError(f"non-finite {field} value {value!r} in {filename!r}")
    return numeric
